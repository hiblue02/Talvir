import requests as rq 
from bs4 import BeautifulSoup as bs
import re
from openpyxl import Workbook

#전체 url 추출
linklist=[]
cn = 0
while cn < 41:
    base_url = 'http://gjicp.ggcf.kr/archives/artworks/si-gun-munhwa1?tid%5B0%5D=36&tid%5B1%5D=37&tid%5B2%5D=38&tid%5B3%5D=39&tid%5B4%5D=40&tid%5B5%5D=41&tid%5B6%5D=42&tid%5B7%5D=43&tid%5B8%5D=44&tid%5B9%5D=45&tid%5B10%5D=46&tid%5B11%5D=47&tid%5B12%5D=33&tid%5B13%5D=48&tid%5B14%5D=49&tid%5B15%5D=50&tid%5B16%5D=35&tid%5B17%5D=51&tid%5B18%5D=52&tid%5B19%5D=53&tid%5B20%5D=54&tid%5B21%5D=123&tid%5B22%5D=34&tid%5B23%5D=56&tid%5B24%5D=57&tid%5B25%5D=58&tid%5B26%5D=59&tid%5B27%5D=60&tid%5B28%5D=61&tid%5B29%5D=62&tid%5B30%5D=63&term=31&pn={codenumber}'.format(codenumber=cn)
    res = rq.get(base_url)
    soup = bs(res.content, 'lxml')

    for link in soup.find_all("div",{"class":"info"}):
        alink=link.find('a')
        try: # 에러발생 시 건너뛰기, none값 발생
            linklist.append(alink.get('href'))
        except:
            pass
    cn +=1

#엑셀삽입
wb = Workbook()
ws = wb.active
ws['A1'] = '명칭'
ws['B1'] = '지정구분'
ws['C1'] = '종목'
ws['D1'] = '유형'
ws['E1'] = '지정일'
ws['F1'] = '소재지'
ws['G1'] = '경도'
ws['H1'] = '위도'
ws['I1'] = '웹자원'
#url별 정보 추출
n=1
for webi in linklist:
    base_url2= webi
    res1 = rq.get(base_url2)
    soup1 = bs(res1.content, 'lxml')

    tablebox = soup1.find_all("td", {"class": "align-l"})
    지정구분 = tablebox[0].text
    종목 = tablebox[1].text
    명칭 = tablebox[2].text
    유형 = tablebox[3].text
    지정일 = tablebox[4].text
    소재지 = tablebox[5].text
    ws[str(chr(65))+str(n+1)] = 명칭
    ws[str(chr(66)) + str(n + 1)] = 지정구분
    ws[str(chr(67)) + str(n + 1)] = 종목
    ws[str(chr(68)) + str(n + 1)] = 유형
    ws[str(chr(69)) + str(n + 1)] = 지정일
    ws[str(chr(70)) + str(n + 1)] = 소재지
    ws[str(chr(73)) + str(n+1)] = webi
    lan = tablebox[5].a.get('href')
    try: # 에러발생 시 건너뛰기, none값 발생
            경도=re.search('x=(.+)\&y', lan).group(1)
            위도 = re.search('y=(.+)\&e', lan).group(1)
            ws[str(chr(71)) + str(n + 1)] = 경도
            ws[str(chr(72)) +str(n+1)] = 위도
    except:
        try:
            경도=re.search('lng=(.+)\&qu',lan).group(1)
            위도 = re.search('lat=(.+)\&ln', lan).group(1)
            ws[str(chr(71)) + str(n + 1)] = 경도
            ws[str(chr(72)) + str(n + 1)] = 위도

        except:
            pass
    n+=1

wb.save('tabletinfo.xlsx')
